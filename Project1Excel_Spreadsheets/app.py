import openpyxl as xl #此处是个 openpyxl 一个 alias(别名)，为了让代码更加整洁。
from openpyxl.chart import BarChart, Reference #从 openpyxl.chart 中导入 BarChart 和 Reference 类 


def process_work(filename):

    wb = xl.load_workbook(f'PythonLearningNote/Project1Excel_Spreadsheets/Excel_file/{filename}') #打开一个 Excel 文件,返回一个 workbook 对象
    sheet = wb['Sheet1'] #返回 Sheet1 对象，存储在 sheet 变量中

    # cell = sheet['a1'] #直接通过square bracket来获取cell
    # cell = sheet.cell(1, 1) #通过 sheet对象的 cell 方法获取单元格对象,第一个参数是行，第二个参数是列


    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.7
        corrected_price_cell = sheet.cell(row, 4)  #sheet.cell会返回一个单元格对象，将其储存在 corrected_price_cell 变量中
        corrected_price_cell.value = corrected_price  #将单元格对象的值设置为 corrected_price

    values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4) #创建一个 Reference 对象，指定数据范围
    #此处的 values 对象有对应范围内表单的所有单元格的值

    chart = BarChart() #创建一个 BarChart 的 instance(实例)
    chart.add_data(values) #将数据添加到chart实例中（图标中）
    sheet.add_chart(chart, 'e2') #将图表添加到工作表中，指定位置为 E2

    wb.save(filename) #保存文件 